from flask import Flask, request, jsonify, send_file, render_template, session
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import re
import time
import io
import os
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = os.urandom(24)

SUAP = "https://suap.ifpi.edu.br"
LOGIN_URL = SUAP + "/accounts/login/"
ADMIN_PROJETOS = SUAP + "/admin/pesquisa/projeto/"

# Armazena sessões por token (em produção usar Redis/DB)
sessions = {}


def get_session(token):
    return sessions.get(token)


def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("X-Session-Token")
        if not token or token not in sessions:
            return jsonify({"error": "Não autenticado"}), 401
        return f(*args, **kwargs)
    return decorated


# --------------------------------------
# LOGIN
# --------------------------------------

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json
    usuario = data.get("usuario", "").strip()
    senha = data.get("senha", "")

    if not usuario or not senha:
        return jsonify({"error": "Usuário e senha são obrigatórios"}), 400

    sess = requests.Session()

    try:
        r = sess.get(LOGIN_URL, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        csrf_input = soup.find("input", {"name": "csrfmiddlewaretoken"})
        if not csrf_input:
            return jsonify({"error": "Não foi possível conectar ao SUAP"}), 502
        csrf = csrf_input["value"]

        payload = {
            "username": usuario,
            "password": senha,
            "csrfmiddlewaretoken": csrf
        }
        headers = {"Referer": LOGIN_URL}
        r = sess.post(LOGIN_URL, data=payload, headers=headers, timeout=15)

        if "logout" not in r.text.lower():
            return jsonify({"error": "Usuário ou senha inválidos"}), 401

        import secrets
        token = secrets.token_hex(32)
        sessions[token] = sess

        return jsonify({"token": token, "usuario": usuario})

    except requests.exceptions.Timeout:
        return jsonify({"error": "Tempo limite de conexão com o SUAP"}), 504
    except Exception as e:
        return jsonify({"error": f"Erro de conexão: {str(e)}"}), 502


# --------------------------------------
# PROJETOS
# --------------------------------------

@app.route("/api/projetos", methods=["GET"])
@require_auth
def get_projetos():
    token = request.headers.get("X-Session-Token")
    sess = get_session(token)
    ano = request.args.get("ano", type=int)

    if not ano or not (2017 <= ano <= 2026):
        return jsonify({"error": "Ano inválido. Informe entre 2017 e 2026"}), 400

    projetos = []
    ids_encontrados = set()
    page = 1

    try:
        while True:
            if page == 1:
                url = f"{ADMIN_PROJETOS}?ano={ano}&tab=tab_em_execucao"
            else:
                url = f"{ADMIN_PROJETOS}?ano={ano}&tab=tab_em_execucao&p={page}"

            r = sess.get(url, timeout=15)
            soup = BeautifulSoup(r.text, "html.parser")
            tabela = soup.find("table", {"id": "result_list"})

            if not tabela:
                break

            linhas = tabela.find("tbody").find_all("tr")
            novos = 0

            for tr in linhas:
                tds = tr.find_all("td")
                if len(tds) < 4:
                    continue

                coordenador = tds[2].get_text(strip=True)
                nome = tds[3].get_text(strip=True)

                link = None
                for a in tr.find_all("a", href=True):
                    if re.match(r"/pesquisa/projeto/\d+/", a["href"]):
                        link = a["href"]
                        break

                if not link:
                    continue

                pid = re.search(r"\d+", link).group()
                if pid in ids_encontrados:
                    continue

                ids_encontrados.add(pid)
                projetos.append({
                    "id": pid,
                    "nome": nome,
                    "coordenador": coordenador,
                    "url": SUAP + link
                })
                novos += 1

            if novos == 0:
                break

            page += 1
            time.sleep(0.5)

        return jsonify({"projetos": projetos, "total": len(projetos)})

    except requests.exceptions.Timeout:
        return jsonify({"error": "Tempo limite ao buscar projetos"}), 504
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --------------------------------------
# DETALHES DE UM PROJETO
# --------------------------------------

def parse_pendencias(soup):
    gerais = []
    conclusao = []
    for p in soup.find_all("p", class_="checklist"):
        texto = p.get_text(" ", strip=True)
        texto = texto.replace("Pendente", "").replace("OK", "").strip()
        if not texto:
            continue
        t = texto.lower()
        if "conclus" in t or "finaliz" in t:
            conclusao.append(texto)
        else:
            gerais.append(texto)
    return gerais, conclusao


def parse_metas(soup):
    metas = []
    metas_pendentes = []
    tabela = soup.find("table")
    if not tabela:
        return metas, metas_pendentes
    for tr in tabela.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue
        atividade = tds[0].get_text(strip=True)
        status = tds[-1].get_text(strip=True)
        metas.append(atividade)
        if "não" in status.lower():
            metas_pendentes.append(atividade)
    return metas, metas_pendentes


@app.route("/api/detalhes/<pid>", methods=["GET"])
@require_auth
def get_detalhes(pid):
    token = request.headers.get("X-Session-Token")
    sess = get_session(token)

    try:
        base = f"{SUAP}/pesquisa/projeto/{pid}/"
        r = sess.get(base + "?tab=pendencias", timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        gerais, conclusao = parse_pendencias(soup)

        r = sess.get(f"{SUAP}/pesquisa/validar_execucao_etapa/{pid}/", timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        metas, metas_pendentes = parse_metas(soup)

        return jsonify({
            "id": pid,
            "pendencias_gerais": gerais,
            "pendencias_conclusao": conclusao,
            "metas": metas,
            "metas_pendentes": metas_pendentes
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --------------------------------------
# GERAR EXCEL
# --------------------------------------

@app.route("/api/exportar", methods=["POST"])
@require_auth
def exportar():
    data = request.json
    projetos = data.get("projetos", [])
    ano = data.get("ano", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Projetos {ano}"

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color="1a5276")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Projeto", "Coordenador",
        "Qtd Pendências Gerais", "Pendências Gerais",
        "Qtd Pendências Conclusão", "Pendências Conclusão",
        "Qtd Metas Pendentes", "Metas Pendentes",
        "URL"
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    for i, p in enumerate(projetos, start=2):
        pend_gerais = "\n".join(p.get("pendencias_gerais", []))
        pend_conc = "\n".join(p.get("pendencias_conclusao", []))
        metas = "\n".join(p.get("metas_pendentes", []))

        row_data = [
            p.get("nome", ""),
            p.get("coordenador", ""),
            len(p.get("pendencias_gerais", [])),
            pend_gerais,
            len(p.get("pendencias_conclusao", [])),
            pend_conc,
            len(p.get("metas_pendentes", [])),
            metas,
            p.get("url", "")
        ]
        ws.append(row_data)

        # Zebra striping
        fill_color = "eaf4fb" if i % 2 == 0 else "FFFFFF"
        row_fill = PatternFill("solid", start_color=fill_color)
        for cell in ws[i]:
            cell.fill = row_fill
            cell.font = Font(name="Arial", size=10)

        for col_letter in ["D", "F", "H"]:
            ws[f"{col_letter}{i}"].alignment = Alignment(wrap_text=True, vertical="top")

    # Larguras
    col_widths = {"A": 45, "B": 28, "C": 12, "D": 70, "E": 12, "F": 70, "G": 12, "H": 55, "I": 50}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome = f"relatorio_suap_{ano}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --------------------------------------
# FRONTEND
# --------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)